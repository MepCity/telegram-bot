"""
Excel dosyasını okuma ve güncelleme modülü
"""
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.styles.colors import Color
from copy import copy
from datetime import datetime
from pathlib import Path
import config
import json


class ExcelHandler:
    """Excel template işlemleri"""
    
    def __init__(self, template_path=None):
        self.template_path = template_path or config.TEMPLATE_PATH
        self.counter_file = Path(config.OUTPUT_DIR) / 'offer_counter.json'
        
    def _get_next_offer_number(self):
        """Benzersiz artan sipariş numarası üret"""
        # Counter dosyasını oku veya oluştur
        if self.counter_file.exists():
            with open(self.counter_file, 'r') as f:
                data = json.load(f)
                counter = data.get('counter', 10000)  # 10000'den başla
        else:
            counter = 10000
            
        # Sayıyı artır
        counter += 1
        
        # Kaydet
        self.counter_file.parent.mkdir(exist_ok=True)
        with open(self.counter_file, 'w') as f:
            json.dump({'counter': counter}, f)
            
        return str(counter)
        
    def create_offer(self, customer_data, services, offer_info):
        """
        Teklif Excel dosyası oluştur - YENİ ÖZEL ŞABLON (yeni.xlsx)
        
        Args:
            customer_data (dict): Müşteri bilgileri
                - name: Müşteri adı
                - contact_person: Yetkili kişi
                - phone: Telefon
                - tax_office: Vergi dairesi
                - tax_number: Vergi no
                - address: Adres
            services (list): Hizmet/ürün listesi
                - name: Hizmet adı
                - unit: Birim (adet, m2, vb)
                - quantity: Miktar
                - unit_price: Birim fiyat
            offer_info (dict): Teklif bilgileri
                - offer_no: Teklif no
                - offer_date: Teklif tarihi
                - subject: Konu
                - validity: Son geçerlilik
                - currency: Para birimi (₺, $, €)
                - discount_rate: İskonto oranı (0-100)
                - notes: Notlar
        
        Returns:
            str: Oluşturulan dosya yolu
        """
        # Template'i yükle (yeni.xlsx - 28.10.2025 güncel şablon)
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active
        
        # Para birimi (varsayılan TL)
        currency = offer_info.get('currency', '₺')
        
        # SATIR 2-3: Tarih ve Teklif No (SAĞ ÜST - K2 ve K3)
        offer_no = offer_info.get('offer_no') or f"T-{datetime.now().strftime('%Y')}-{self._get_next_offer_number()}"
        
        # helper: merged hücreleri yöneten güvenli set
        def set_cell(r, c, v):
            """Row, Col (1-index) ile yazma; eğer hücre merged ise top-left hücreye yaz."""
            cell = ws.cell(row=r, column=c)
            try:
                cell.value = v
                return
            except AttributeError:
                # MergedCell ise, içeren aralığı bulup top-left'e yaz
                for merged in ws.merged_cells.ranges:
                    if merged.min_row <= r <= merged.max_row and merged.min_col <= c <= merged.max_col:
                        ws.cell(row=merged.min_row, column=merged.min_col).value = v
                        return
                # fallback
                ws.cell(row=r, column=c).value = v
        
        def set_cell_with_style(r, c, v):
            """Hücreye yazarken mevcut font ve stil bilgilerini koru"""
            cell = ws.cell(row=r, column=c)
            
            # Mevcut stil bilgilerini kaydet
            old_font = copy(cell.font) if cell.font else None
            old_alignment = copy(cell.alignment) if cell.alignment else None
            old_fill = copy(cell.fill) if cell.fill else None
            old_border = copy(cell.border) if cell.border else None
            old_number_format = cell.number_format
            
            try:
                # Değeri yaz
                cell.value = v
                
                # Eski stilleri geri yükle
                if old_font:
                    cell.font = old_font
                if old_alignment:
                    cell.alignment = old_alignment
                if old_fill:
                    cell.fill = old_fill
                if old_border:
                    cell.border = old_border
                if old_number_format:
                    cell.number_format = old_number_format
                    
            except AttributeError:
                # MergedCell ise, içeren aralığı bulup top-left'e yaz
                for merged in ws.merged_cells.ranges:
                    if merged.min_row <= r <= merged.max_row and merged.min_col <= c <= merged.max_col:
                        target_cell = ws.cell(row=merged.min_row, column=merged.min_col)
                        
                        # Hedef hücrenin stilini kopyala
                        old_font = copy(target_cell.font) if target_cell.font else None
                        old_alignment = copy(target_cell.alignment) if target_cell.alignment else None
                        old_fill = copy(target_cell.fill) if target_cell.fill else None
                        old_border = copy(target_cell.border) if target_cell.border else None
                        old_number_format = target_cell.number_format
                        
                        target_cell.value = v
                        
                        # Stilleri geri yükle
                        if old_font:
                            target_cell.font = old_font
                        if old_alignment:
                            target_cell.alignment = old_alignment
                        if old_fill:
                            target_cell.fill = old_fill
                        if old_border:
                            target_cell.border = old_border
                        if old_number_format:
                            target_cell.number_format = old_number_format
                        return
                
                # fallback
                ws.cell(row=r, column=c).value = v
        
        # FİX: Tablo başlıklarının (Satır 14) ve toplam satırlarının (Satır 22) font renklerini beyaz yap
        # Koyu arka planda (FF34495D) siyah yazılar görünmüyor, beyaz olmalı
        white_font_cells = ['A14', 'H14', 'I14', 'J14', 'G22', 'K22']  # C14 zaten beyaz şablonda
        for cell_ref in white_font_cells:
            cell = ws[cell_ref]
            if cell.font:
                # Mevcut font'u kopyala ama rengi beyaz yap
                new_font = copy(cell.font)
                new_font.color = Color(rgb='FFFFFFFF')
                cell.font = new_font

        # K2: Tarih
        set_cell(2, 11, f"Tarih: {offer_info.get('offer_date', datetime.now().strftime('%d.%m.%Y'))}")
        
        # K3: Teklif No
        set_cell(3, 11, f"Teklif No: {offer_no}")
        
        # SATIR 9-12: MÜŞTERİ BİLGİLERİ (Sol taraf - C sütunu)
        set_cell(9, 3, customer_data.get('name', ''))  # C9 = Firma Adı
        set_cell(10, 3, customer_data.get('contact_person', ''))  # C10 = Yetkili
        set_cell(11, 3, customer_data.get('phone', ''))  # C11 = Telefon
        
        # SATIR 9-12: TEKLİF BİLGİLERİ (Sağ taraf - H sütunu)
        set_cell(9, 8, offer_no)  # H9 = Teklif Sipariş No
        set_cell(10, 8, offer_info.get('offer_date', datetime.now().strftime('%d.%m.%Y')))  # H10 = Teklif Tarihi
        set_cell(11, 8, 'PEŞİN')  # H11 = Ödeme Şekli (her zaman PEŞİN)
        set_cell(12, 8, offer_info.get('delivery_date', ''))  # H12 = Planlanan Teslim Tarihi
        
        # SATIR 16+: ÜRÜN TABLOSU
        # Şablonda 16, 17, 18 satırları hazır - önce temizle
        for row in range(16, 19):
            set_cell(row, 1, None)  # A: NO
            set_cell(row, 3, None)  # C: HİZMET (C-D merged)
            set_cell(row, 8, None)  # H: MİKTAR
            set_cell(row, 9, None)  # I: BİRİM FİYAT
            set_cell(row, 10, None)  # J: TUTAR (J-K merged)
        
        # Yeni ürünleri ekle
        start_row = 16
        total_amount = 0
        
        for idx, service in enumerate(services):
            if idx >= 3:  # Maksimum 3 ürün (şablonda 16, 17, 18)
                break
                
            row = start_row + idx
            quantity = service.get('quantity', 1)
            unit_price = service.get('unit_price', 0)
            amount = quantity * unit_price
            total_amount += amount
            
            # A: NO
            set_cell(row, 1, idx + 1)

            # C: HİZMET AÇIKLAMASI (C-D merged)
            set_cell(row, 3, service.get('name', ''))

            # H: MİKTAR
            set_cell(row, 8, quantity)

            # I: BİRİM FİYAT
            set_cell(row, 9, f"{unit_price:,.2f} {currency}")

            # J: TUTAR (J-K merged)
            set_cell(row, 10, f"{amount:,.2f} {currency}")
        
        # SATIR 20-22: FİYAT HESAPLAMALARI (font renklerini koru)
        # K20: Ara Toplam
        set_cell_with_style(20, 11, f"{total_amount:,.2f} {currency}")
        
        # K21: KDV (%25)
        kdv_amount = total_amount * config.KDV_RATE
        set_cell_with_style(21, 11, f"{kdv_amount:,.2f} {currency}")
        
        # K22: Genel Toplam (K22-K23 merged, mavi arka plan)
        grand_total = total_amount + kdv_amount
        set_cell_with_style(22, 11, f"{grand_total:,.2f} {currency}")
        
        # SATIR 20: NOTLAR (A20 hücresi - isteğe bağlı)
        # Şablondaki varsayılan notu temizle ve kullanıcının notunu ekle
        notes = offer_info.get('notes', '')
        if notes:
            set_cell(20, 1, notes)  # A20'ye kullanıcının notunu yaz
        else:
            set_cell(20, 1, '')  # Boş bırak (şablondaki varsayılan notu sil)
        
        # SATIR 26: Sipariş Alan / Teklif Veren
        set_cell(26, 1, "Adı Soyadı: Hatice Arslan")
        
        # SATIR 25: İmza satırı (şablonda zaten "İmza" yazıyor - dokunma)
        
        # Page setup: A4 dikey (portrait) tek sayfaya sığdırma
        try:
            ps = ws.page_setup
            ps.fitToPage = True
            ps.fitToHeight = 1
            ps.fitToWidth = 1
            ps.orientation = 'portrait'  # A4 dikey
            ps.paperSize = 9  # A4
            
            # Dar margin ile sığdırmayı kolaylaştır
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            
            # Print area: A1'den K28'e kadar (yeni.xlsx aralığı - 2 satır eklendi)
            ws.print_area = "A1:K28"
        except Exception as e:
            # Hata olursa es geç
            pass
        
        # Çıktı dizinini oluştur
        output_dir = Path(config.OUTPUT_DIR)
        output_dir.mkdir(exist_ok=True)
        
        # Dosya adı oluştur
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        customer_name_safe = customer_data.get('name', 'musteri').replace(' ', '_')[:30]
        output_path = output_dir / f'teklif_{customer_name_safe}_{timestamp}.xlsx'
        
        # Kaydet
        wb.save(output_path)
        
        return str(output_path)
    
    def _generate_offer_no(self):
        """DEPRECATED: Kullanılmıyor, _get_next_offer_number() kullanın"""
        return self._get_next_offer_number()


if __name__ == '__main__':
    # Test
    handler = ExcelHandler()
    
    test_customer = {
        'name': 'TEST FİRMA A.Ş.',
        'contact_person': 'AHMET BEY',
    }
    
    test_services = [
        {
            'name': 'YATIRIM TEŞVİK BELGESİ',
            'quantity': 1,
            'unit_price': 750000
        }
    ]
    
    test_offer_info = {
        'offer_no': '05274',
        'offer_date': '28.10.2025',
        'payment_method': 'PEŞİN',
        'delivery_date': '--------------------'
    }
    
    output = handler.create_offer(test_customer, test_services, test_offer_info)
    print(f'Test teklif oluşturuldu: {output}')
